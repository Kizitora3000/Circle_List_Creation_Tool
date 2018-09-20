import openpyxl
from openpyxl.styles import Alignment
import sys


def input_to_excel(inputted_string, title):
    #デバック用
    #print("Input_to_Excel")
    wb = openpyxl.load_workbook(title + ".xlsx")
    sheet = wb.active

    "#入力したいセルの場所を取得"
    try:
        max_row = sheet.max_row
        max_row += 1

        circle_cell = sheet.cell(row=max_row, column=1)
        things_cell = sheet.cell(row=max_row, column=4)
        twitter_address_cell = sheet.cell(row=max_row, column=9)
        priority_cell = sheet.cell(row=max_row, column=11)
        how_much_cell = sheet.cell(row=max_row, column=12)
        supplement_cell = sheet.cell(row=max_row, column=14)

        "#入力したいセルの場所を結合"
        sheet.merge_cells(start_row=max_row, start_column=1, end_row=max_row, end_column=3)
        sheet.merge_cells(start_row=max_row, start_column=4, end_row=max_row, end_column=8)
        sheet.merge_cells(start_row=max_row, start_column=9, end_row=max_row, end_column=10)
        sheet.merge_cells(start_row=max_row, start_column=12, end_row=max_row, end_column=13)
        sheet.merge_cells(start_row=max_row, start_column=14, end_row=max_row, end_column=16)

        "#文字揃えの変更（すべて中央揃えにする）"
        "#変更した値をリストに入れ直す"
        cell_list = []

        cell_list.extend([circle_cell,
                          things_cell,
                          twitter_address_cell,
                          priority_cell,
                          how_much_cell,
                          supplement_cell,
                          ])

        for cell in cell_list:
            cell.alignment = Alignment(horizontal='center', vertical='center')

        "#入力したいセルの行のサイズを変更"
        sheet.row_dimensions[max_row].height = 18.00
    except NameError:
        sys.exit()

    "#文字列の入力"
    for i in range(6):
        cell_list[i].value = inputted_string[i]

    wb.save(title + ".xlsx")
