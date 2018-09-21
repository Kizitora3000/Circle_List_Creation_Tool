import sys
import openpyxl
from openpyxl.styles.borders import Side, Border


def quit_excel(title):
    # デバック用
    # print("quit_Excel.py")

    wb = openpyxl.load_workbook(title + ".xlsx")
    sheet = wb.active

    def draw_border_to_cell(ws, top_left_cell, bottom_right_cell, border_pattern):
        for _row in ws[top_left_cell:bottom_right_cell]:
            for _cell in _row:
                _cell.border = border_pattern

    border = Border(
        left=Side(color='FF000000', style='thin'),
        right=Side(color='FF000000', style='thin'),
        top=Side(color='FF000000', style='thin'),
        bottom=Side(color='FF000000', style='thin')
    )

    # 最も右下のセルを探す（罫線を引くときに使う）
    max_row = sheet.max_row

    top_left_cell = sheet.cell(row=1, column=1)
    bottom_right_cell = sheet.cell(row=max_row, column=16)

    top_left_cell = top_left_cell.coordinate
    bottom_right_cell = bottom_right_cell.coordinate
    print(top_left_cell)
    print(bottom_right_cell)

    draw_border_to_cell(sheet, top_left_cell, bottom_right_cell, border)

    wb.save(title + ".xlsx")

    sys.exit(
