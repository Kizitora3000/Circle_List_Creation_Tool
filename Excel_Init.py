import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import Font


def excel_init(title):
    #デバック用
    #print("Excel_Init.py")
    wb = openpyxl.Workbook()
    sheet = wb.active

    "#セルの結合"
    merge_cells = ["A1:P2", "A3:C3", "D3:H3", "I3:J3", "L3:M3", "N3:P3"]
    for merge_cell in merge_cells:
        sheet.merge_cells(merge_cell)

    "#セルのサイズ変更"
    sheet.row_dimensions[1].height = 18.00
    sheet.row_dimensions[2].height = 18.00
    sheet.row_dimensions[3].height = 26.40

    "#セルの揃え変更"
    cell_list = []

    title_cell = sheet.cell(row=1, column=1)
    circle_cell = sheet.cell(row=3, column=1)
    things_cell = sheet.cell(row=3, column=4)
    twitter_address_cell = sheet.cell(row=3, column=9)
    priority_cell = sheet.cell(row=3, column=11)
    how_much_cell = sheet.cell(row=3, column=12)
    supplement_cell = sheet.cell(row=3, column=14)

    "#文字揃えの変更（すべて中央揃えにする）"
    cell_list.extend([title_cell,
                      circle_cell,
                      things_cell,
                      twitter_address_cell,
                      priority_cell,
                      how_much_cell,
                      supplement_cell,
                      ])

    for cell in cell_list:
        cell.alignment = Alignment(horizontal='center', vertical='center')

    "#フォント"
    font_title = Font(name="游ゴシック", size=24, bold=True)
    font_twitter_address = Font(name="游ゴシック", size=14)
    font_priority = Font(name="游ゴシック", size=12)
    font_others = Font(name="游ゴシック", size=16)

    "#フォントの設定付与と文字入力"
    cell_list[0].font = font_title
    cell_list[0].value = title

    circle_cell.font = font_others
    circle_cell.value = "サークル名"

    things_cell.font = font_others
    things_cell.value = "買う内容"

    twitter_address_cell.font = font_twitter_address
    twitter_address_cell.value = "twitterアドレス"

    priority_cell.font = font_priority
    priority_cell.value = "優先順位"

    how_much_cell.font = font_others
    how_much_cell.value = "予想金額"

    supplement_cell.font = font_others
    supplement_cell.value = "補足"

    wb.save(title + ".xlsx")
